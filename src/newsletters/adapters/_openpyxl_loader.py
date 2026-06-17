"""Lazy openpyxl boundary — the optional ``[excel]`` extra, imported only on use (ADAPT-03).

WHY THIS MODULE EXISTS (CONTEXT decision 1 / R3, threat T-05-04 + T-05-05):
``openpyxl`` is a third-party dependency that backs the Excel adapter. It is NOT AI, so the
``forbid-ai-in-core`` import-linter contract is unaffected — but the AI-optional / minimal-core
invariant still demands that a bare ``pip install .`` (no ``[excel]``) can ``import newsletters``,
``import newsletters.adapters``, and run the deterministic spine with zero openpyxl. We achieve
that by mirroring the ``[ai]`` lazy-import discipline: **openpyxl is never imported at module
top-level anywhere reachable from ``import newsletters``.** It is imported INSIDE
:func:`_load_openpyxl`, which the Excel adapter (Plan 05-03) calls only when it actually parses a
workbook. Absence raises a clear teaching :class:`ImportError` pointing at ``pip install '.[excel]'``.

This module itself must therefore have NO top-level ``import openpyxl`` / ``from openpyxl ...`` —
the bare-install gate (``tests/test_ai_optional.py``) asserts grep-count 0 for those edges.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING, Any, Union

if TYPE_CHECKING:  # pragma: no cover - typing only; never executed at runtime
    # Under TYPE_CHECKING this import is invisible to the runtime interpreter (and to the
    # bare-install gate's grep for a top-level runtime ``import openpyxl``), so it costs the
    # bare install nothing while still giving mypy the real types.
    from openpyxl import Workbook

# The exact teaching message a user sees if they reach the Excel adapter without the extra.
# Exposed as a module constant so tests can assert against it without string-duplication drift.
MISSING_OPENPYXL_MESSAGE = (
    "The Excel adapter requires the optional 'openpyxl' dependency. "
    "Install it with: pip install '.[excel]'  (or: pip install newsletters[excel]). "
    "The deterministic spine runs without it — openpyxl is needed only for .xlsx extraction "
    "(AI-optional / minimal-core: third-party adapter deps live behind extras)."
)


def _load_openpyxl() -> Any:
    """Import and return the ``openpyxl`` module, lazily.

    The import lives INSIDE this function (never at module top) so that importing
    ``newsletters`` / ``newsletters.adapters`` never requires the ``[excel]`` extra. If openpyxl
    is not installed, re-raise a teaching :class:`ImportError` naming the extra and the install
    command, and stating the spine runs without it.

    Returns:
        The imported ``openpyxl`` module.

    Raises:
        ImportError: if ``openpyxl`` is not installed, with an actionable message.
    """
    try:
        import openpyxl  # noqa: PLC0415 — deliberately lazy (optional [excel] extra, T-05-04)
    except ImportError as exc:
        raise ImportError(MISSING_OPENPYXL_MESSAGE) from exc
    return openpyxl


def load_workbook_pair(
    path_or_bytes: Union[str, bytes, "io.BytesIO"],
) -> "tuple[Workbook, Workbook]":
    """Load one ``.xlsx`` source twice — the formula view and the data view (R3 double-load).

    Faithful Excel extraction requires distinguishing a *genuinely blank* cell from a *formula
    cell whose cached value is ``None``* (an openpyxl-saved file never writes a cache, so
    ``data_only=True`` returns ``None`` for such a formula). Resolving that requires BOTH views:

    - **formula view** (``data_only=False``): ``cell.data_type == 'f'`` identifies formula cells;
      ``cell.value`` is the formula string / the literal value for non-formula cells.
    - **data view** (``data_only=True``): ``cell.value`` is the cached computed value, or ``None``
      if Excel never calculated/saved it.

    Both are loaded with ``read_only=False`` (standard mode) — REQUIRED for faithful merged-cell
    and chart/image accounting (in ``read_only=True`` merged cells become ``<EmptyCell>`` and the
    ``_charts``/``_images`` detection attributes are absent entirely). ``keep_links=False`` avoids
    resolving external-workbook caches.

    The Excel adapter (Plan 05-03) consumes this pair by zipping the two views cell-by-cell.

    Args:
        path_or_bytes: a filesystem path (``str``), raw ``.xlsx`` bytes, or a ``BytesIO``. Bytes
            are wrapped in a fresh ``BytesIO`` per view (a stream is consumed by the first load,
            so each view needs its own buffer).

    Returns:
        ``(workbook_formula_view, workbook_data_view)`` — both standard-mode workbooks. The caller
        is responsible for ``.close()``-ing each when done.

    Raises:
        ImportError: if ``openpyxl`` is not installed (via :func:`_load_openpyxl`).
    """
    openpyxl = _load_openpyxl()

    def _source_for_view() -> "Union[str, io.BytesIO]":
        # A path can be reopened by openpyxl per call; raw bytes / a BytesIO are single-use
        # streams, so each view gets its OWN BytesIO over the same bytes.
        if isinstance(path_or_bytes, (bytes, bytearray)):
            return io.BytesIO(bytes(path_or_bytes))
        if isinstance(path_or_bytes, io.BytesIO):
            return io.BytesIO(path_or_bytes.getvalue())
        return path_or_bytes  # str path

    wb_formula = openpyxl.load_workbook(
        _source_for_view(), data_only=False, read_only=False, keep_links=False
    )
    wb_data = openpyxl.load_workbook(
        _source_for_view(), data_only=True, read_only=False, keep_links=False
    )
    return wb_formula, wb_data


__all__ = ["_load_openpyxl", "load_workbook_pair", "MISSING_OPENPYXL_MESSAGE"]
