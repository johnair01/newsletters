"""Round-trip coverage-parity conformance test (TASK ZERO — the headline proof).

This file is the executable proof that the Phase-4 flaw (04-VERIFICATION.md, the documented
LIMITATION) is fixed: an adapter's ``unextracted[]`` determination now travels WITH the
``Source`` (the typed ``Source.extraction`` carrier, R1), so re-``distill()``ing a *persisted*
``Source`` on a FRESH adapter reproduces the original coverage — never a silent drop, never a
false ``complete=True``.

It has three layers:

1. **Carrier + codec unit tests** (Task 1) — ``ExtractedDrop`` / ``ExtractionRecord`` round-trip
   losslessly via JSON; ``Source.extraction`` defaults to ``None`` (backward-compatible) and
   round-trips; the shared ``encode_coverage`` / ``decode_coverage`` codec is a lossless,
   order-preserving, total bridge between ``Unextracted`` and ``ExtractionRecord``.

2. **Round-trip parity** (Task 2, the headline) — for every committed fixture: parse on adapter
   A, serialize+reload the Source (``model_dump_json -> model_validate_json``), then distill on a
   FRESH adapter B (B never saw ``parse()``). B's coverage EQUALS A's coverage. Parametrized over
   the registered adapters so the Excel adapter (Plan 03) joins the same matrix.

3. **R2 safety-net** (Task 2) — ``distill()`` handed a Source with NO reconstructable coverage
   (not produced by this adapter AND no ``extraction`` record) reports ``complete=False`` with an
   explicit ``coverage-not-reconstructable`` marker — never a false ``complete=True``.
"""

from __future__ import annotations

import importlib.util
import io
import pathlib

import pytest

from newsletters.adapters._coverage_codec import (
    COVERAGE_NOT_RECONSTRUCTABLE,
    decode_coverage,
    encode_coverage,
)
from newsletters.adapters.email_adapter import EmailAdapter
from newsletters.distill.coverage import Unextracted
from newsletters.locators import (
    ExtractedDrop,
    ExtractionRecord,
    FreeLocator,
    SessionLocator,
)
from newsletters.semantic import Source

# --------------------------------------------------------------------------- #
# A parametrizable adapter matrix. Each entry is a self-contained recipe so the
# Excel adapter (Plan 03) joins by appending one entry — no test-body change.
# Each adapter case provides:
#   - factory():            a FRESH adapter instance
#   - fixtures():           a list of (name, raw_bytes, path) to parse
# --------------------------------------------------------------------------- #

_EML_DIR = pathlib.Path(__file__).parent / "fixtures" / "eml"


def _email_fixtures() -> list[tuple[str, bytes, str]]:
    return [
        (p.name, p.read_bytes(), str(p))
        for p in sorted(_EML_DIR.glob("*.eml"))
    ]


# The Excel adapter joins the SAME matrix (Plan 03). openpyxl is the optional [excel] extra, so the
# whole Excel case is skipped cleanly when it is absent — the bare-install gate is unaffected. The
# committed byte-reproducible golden corpus is Plan 04; here we author tiny in-memory .xlsx bytes
# covering a representative spread (literal, formula-cache gap, error cell) so coverage parity is
# proven for Excel too.
_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None


def _xlsx_fixtures() -> list[tuple[str, bytes, str]]:
    import openpyxl

    def _bytes(build) -> bytes:
        wb = openpyxl.Workbook()
        build(wb.active)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _literals(ws) -> None:
        ws.title = "S"
        ws["A1"] = "text"
        ws["A2"] = 42

    def _formula_gap(ws) -> None:
        ws.title = "F"
        ws["A1"] = 2
        ws["A2"] = "=A1*10"  # openpyxl writes no cache -> a formula-cache-gap drop

    def _error_cell(ws) -> None:
        ws.title = "E"
        ws["A1"] = "#DIV/0!"  # an error cell -> a drop

    return [
        ("literals.xlsx", _bytes(_literals), "literals.xlsx"),
        ("formula_gap.xlsx", _bytes(_formula_gap), "formula_gap.xlsx"),
        ("error_cell.xlsx", _bytes(_error_cell), "error_cell.xlsx"),
    ]


def _excel_adapter_factory():
    from newsletters.adapters.excel_adapter import ExcelAdapter

    return ExcelAdapter()


# The PowerPoint adapter (Plan 06-03) joins the SAME matrix. python-pptx is the optional [pptx]
# extra, so the whole pptx case is skipped cleanly when it is absent — the bare-install gate is
# unaffected. Unlike the in-memory excel recipe above, the pptx case reads the COMMITTED
# byte-reproducible golden corpus (Plan 06-04, tests/fixtures/pptx/*.pptx) so parity is proven over
# the real fixtures including the SmartArt + nested-group decks.
_HAS_PPTX = importlib.util.find_spec("pptx") is not None
_PPTX_DIR = pathlib.Path(__file__).parent / "fixtures" / "pptx"


def _pptx_fixtures() -> list[tuple[str, bytes, str]]:
    return [(p.name, p.read_bytes(), str(p)) for p in sorted(_PPTX_DIR.glob("*.pptx"))]


def _pptx_adapter_factory():
    from newsletters.adapters.pptx_adapter import PptxAdapter

    return PptxAdapter()


# (id, adapter_factory, fixtures_loader)
ADAPTER_CASES = [
    pytest.param(EmailAdapter, _email_fixtures, id="email"),
    pytest.param(
        _excel_adapter_factory,
        _xlsx_fixtures,
        id="excel",
        marks=pytest.mark.skipif(
            not _HAS_OPENPYXL, reason="optional [excel] extra (openpyxl) not installed"
        ),
    ),
    pytest.param(
        _pptx_adapter_factory,
        _pptx_fixtures,
        id="pptx",
        marks=pytest.mark.skipif(
            not _HAS_PPTX, reason="optional [pptx] extra (python-pptx) not installed"
        ),
    ),
]


# A small, representative spread of Unextracted drops covering both locator variants
# and an empty reason, used by the codec round-trip identity test.
_SAMPLE_DROPS = [
    Unextracted(locator=FreeLocator(text="Sheet1!A1"), reason="formula cache None"),
    Unextracted(locator=FreeLocator(text=""), reason=""),
    Unextracted(
        locator=SessionLocator(source_id="s-7", note="attachment"),
        reason="non-text attachment — not extracted",
    ),
]


# ============================================================================ #
# Task 1 — carrier + codec
# ============================================================================ #


def test_extracted_drop_roundtrips_via_json() -> None:
    """An ExtractedDrop round-trips losslessly through model_dump_json -> model_validate_json."""
    drop = ExtractedDrop(locator=FreeLocator(text="A1"), reason="because")
    assert ExtractedDrop.model_validate_json(drop.model_dump_json()) == drop


def test_extraction_record_roundtrips_via_json() -> None:
    """An ExtractionRecord with several drops round-trips losslessly through JSON."""
    record = ExtractionRecord(
        unextracted=[
            ExtractedDrop(locator=FreeLocator(text="A1"), reason="r1"),
            ExtractedDrop(
                locator=SessionLocator(source_id="s1", note="n"), reason="r2"
            ),
        ]
    )
    assert ExtractionRecord.model_validate_json(record.model_dump_json()) == record


def test_extraction_record_defaults_empty() -> None:
    """An ExtractionRecord constructed with no args has an empty unextracted[] (total default)."""
    assert ExtractionRecord().unextracted == []


def test_source_extraction_defaults_none_and_is_backward_compatible() -> None:
    """A Source built without `extraction` has None (backward-compatible default).

    Crucially, an existing Rev1/Phase-4 Source JSON (which has NO `extraction` key at all) must
    still validate — the field is optional with a None default.
    """
    s = Source(id="x", transcript="hi")
    assert s.extraction is None
    # legacy JSON: no `extraction` key whatsoever still loads
    legacy = '{"id": "x", "transcript": "hi", "context": ""}'
    assert Source.model_validate_json(legacy).extraction is None


def test_source_with_extraction_roundtrips() -> None:
    """A Source carrying an ExtractionRecord round-trips losslessly through JSON."""
    s = Source(
        id="x",
        transcript="hi",
        extraction=ExtractionRecord(
            unextracted=[ExtractedDrop(locator=FreeLocator(text="A1"), reason="r")]
        ),
    )
    reloaded = Source.model_validate_json(s.model_dump_json())
    assert reloaded == s
    assert reloaded.extraction == s.extraction


def test_extraction_excluded_from_content_hash() -> None:
    """The coverage carrier is metadata about extraction, NOT the addressed content.

    Setting `extraction` must NOT change `content_hash()` — every existing Trace stays
    addressed and non-stale across the carrier's introduction.
    """
    bare = Source(id="x", transcript="hi")
    carried = Source(
        id="x",
        transcript="hi",
        extraction=ExtractionRecord(
            unextracted=[ExtractedDrop(locator=FreeLocator(text="A1"), reason="r")]
        ),
    )
    assert bare.content_hash() == carried.content_hash()


def test_encode_coverage_preserves_locator_and_reason_in_order() -> None:
    """encode_coverage maps each Unextracted to an ExtractedDrop verbatim, order-preserving."""
    record = encode_coverage(_SAMPLE_DROPS)
    assert isinstance(record, ExtractionRecord)
    assert len(record.unextracted) == len(_SAMPLE_DROPS)
    for drop, original in zip(record.unextracted, _SAMPLE_DROPS):
        assert drop.locator == original.locator
        assert drop.reason == original.reason


def test_decode_coverage_none_returns_empty() -> None:
    """A Source not produced by an adapter has no carrier; decode(None) -> [] (total, no raise)."""
    assert decode_coverage(None) == []


def test_codec_roundtrip_identity() -> None:
    """decode_coverage(encode_coverage(drops)) == drops for arbitrary Unextracted (both variants)."""
    assert decode_coverage(encode_coverage(_SAMPLE_DROPS)) == _SAMPLE_DROPS


def test_codec_roundtrip_identity_empty() -> None:
    """An empty drop list round-trips to an empty drop list."""
    assert decode_coverage(encode_coverage([])) == []


def test_codec_survives_source_json_roundtrip() -> None:
    """The whole point: encode onto a Source, JSON round-trip the Source, decode -> identical."""
    s = Source(id="x", transcript="hi", extraction=encode_coverage(_SAMPLE_DROPS))
    reloaded = Source.model_validate_json(s.model_dump_json())
    assert decode_coverage(reloaded.extraction) == _SAMPLE_DROPS


# ============================================================================ #
# Task 2 — round-trip parity (the headline) + R2 safety-net
# ============================================================================ #


def _coverage_signature(coverage: object) -> tuple[bool, list[tuple[str, str]]]:
    """A comparable signature of a Coverage: (complete, [(locator.display, reason), ...]).

    Compares the honest content (the complete flag + the ordered drop identities) without
    depending on object identity, so adapter A's coverage and a fresh adapter B's coverage can
    be asserted EQUAL across a Source round-trip.
    """
    return (
        coverage.complete,  # type: ignore[attr-defined]
        [(u.locator.display, u.reason) for u in coverage.unextracted],  # type: ignore[attr-defined]
    )


@pytest.mark.parametrize("factory,fixtures", ADAPTER_CASES)
def test_roundtrip_coverage_parity(factory, fixtures) -> None:
    """THE headline: persist a Source, distill it on a FRESH adapter -> coverage UNCHANGED.

    Today (pre-fix) this FAILS for the Email adapter: fresh adapter B has an empty in-memory dict,
    so it drops U1-U7 and falsely reports complete=True. With the carrier, coverage is a pure
    function of the (round-tripped) Source.
    """
    for name, raw, path in fixtures():
        adapter_a = factory()
        source, _units, _drops = adapter_a.parse(raw, path)
        original = adapter_a.distill([source])

        # persist + reload (the round-trip boundary), then a FRESH adapter that NEVER parsed it
        reloaded = Source.model_validate_json(source.model_dump_json())
        adapter_b = factory()
        replayed = adapter_b.distill([reloaded])

        assert _coverage_signature(replayed.coverage) == _coverage_signature(
            original.coverage
        ), f"{name}: coverage drifted across Source round-trip on a fresh adapter"


@pytest.mark.parametrize("factory,fixtures", ADAPTER_CASES)
def test_roundtrip_parity_marker_absent_for_real_fixtures(factory, fixtures) -> None:
    """For real fixtures coverage IS reconstructable via the carrier, so R2's marker NEVER fires."""
    for name, raw, path in fixtures():
        adapter_a = factory()
        source, _units, _drops = adapter_a.parse(raw, path)
        reloaded = Source.model_validate_json(source.model_dump_json())
        replayed = factory().distill([reloaded])
        assert all(
            u.reason != COVERAGE_NOT_RECONSTRUCTABLE
            for u in replayed.coverage.unextracted
        ), f"{name}: the coverage-not-reconstructable marker fired for a reconstructable Source"


@pytest.mark.parametrize("factory,fixtures", ADAPTER_CASES)
def test_r2_safety_net_unaccountable_source(factory, fixtures) -> None:
    """R2: a Source with NO carrier, not produced by this adapter, never reports complete=True.

    It MUST report complete=False with an explicit 'coverage-not-reconstructable' marker. Honest
    uncertainty over silent completeness — belt-and-suspenders even when the carrier is absent.
    """
    # a hand-built Source with content but NO extraction carrier and never parse()d by the adapter
    orphan = Source(id="orphan-1", context="hand-built", transcript="some text here")
    assert orphan.extraction is None
    result = factory().distill([orphan])
    assert result.coverage.complete is False, (
        "an unaccountable Source must NOT report complete=True (R2 safety-net)"
    )
    assert any(
        u.reason == COVERAGE_NOT_RECONSTRUCTABLE
        for u in result.coverage.unextracted
    ), "the explicit coverage-not-reconstructable marker must be present"
