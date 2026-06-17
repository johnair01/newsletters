"""Author the byte-reproducible golden `.xlsx` corpus (ADAPT-06; CONTEXT decision 6).

These fixtures stand in for untrusted real-world spreadsheets. Each is tiny, programmatic, and
uses EXPLICIT cell values/types so the Excel adapter's extract-vs-disclose fork is exercised, not
incidental. Mirrors `tests/fixtures/eml/_author_fixtures.py`: the committed `.xlsx` files ARE the
corpus; this script is the documented, rerunnable provenance and lets anyone regenerate them
byte-for-byte.

BYTE-REPRODUCIBILITY (threat T-05-11). openpyxl drifts the saved bytes run-to-run in TWO ways,
both of which we pin:
  1. It stamps the current wall-clock into `docProps/core.xml` (`created`/`modified`) — pinned by
     setting `wb.properties.created`/`modified` to a FIXED `datetime`.
  2. An `.xlsx` is a ZIP, and openpyxl writes each ZIP entry's local-header `date_time` from
     `datetime.now()` at save time (independent of the doc properties) — pinned by `_normalize_zip`,
     which rewrites every entry's `date_time` to a fixed constant and re-deflates in entry order.
With both pinned the generator is fully deterministic (no `now()`, no `random`), so `sha256(file)`
is stable across processes.

THE FORMULA-CACHE CRUX (the faithfulness fork, ADAPT-03 criterion 2). openpyxl never COMPUTES a
formula, so an openpyxl-saved formula cell carries NO cached value — `data_only=True` reads `None`.
That is the natural `formula_no_cache` case: the adapter must route it to `unextracted[]`, NEVER
emit `0`/`""`. To author the contrasting `formula_with_cache` case we inject a `<v>` cached value
directly into the worksheet XML after save (openpyxl has no API to write a formula cache), so the
DATA view reads a real value and the adapter emits it as a verbatim claim. The XML-injection
technique is documented inline at `_with_cache_bytes`.

THE SILENT-LOSS PATH (charts/images). The adapter discloses workbook-level objects it can read but
does not extract (`ws._charts` / `ws._images`). We author a CHART (Pillow is not a declared
dependency, so a real image cannot be embedded; a chart exercises the identical disclosure path via
`ws._charts`). The chart's CONTENT is never extracted — only its presence is disclosed.

Run:  .venv/bin/python tests/fixtures/xlsx/_author_fixtures.py
"""

from __future__ import annotations

import io
import pathlib
import zipfile
from datetime import date, datetime

import openpyxl
from openpyxl.chart import BarChart, Reference

HERE = pathlib.Path(__file__).parent

# A fixed timestamp pinned into docProps/core.xml so the OOXML bytes do not embed a wall-clock
# (byte-reproducibility, threat T-05-11). Any fixed instant works; this one is arbitrary + stable.
_FIXED = datetime(2026, 1, 1, 0, 0, 0)
# The fixed ZIP local-header date_time (year, month, day, hour, minute, second) for every entry.
_FIXED_ZIP_DATE_TIME = (2026, 1, 1, 0, 0, 0)


def _normalize_zip(raw: bytes) -> bytes:
    """Rewrite every ZIP entry's date_time to a fixed constant, preserving entry order + content.

    openpyxl stamps each entry's local-header timestamp from the save-time wall-clock, so two saves
    of the SAME workbook differ only in those embedded times. We rebuild the archive copying each
    member's bytes verbatim but forcing a fixed `date_time`, making the `.xlsx` byte-reproducible.
    """
    zin = zipfile.ZipFile(io.BytesIO(raw))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():  # original entry order -> deterministic
            data = zin.read(item.filename)
            info = zipfile.ZipInfo(item.filename, date_time=_FIXED_ZIP_DATE_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = item.external_attr
            zout.writestr(info, data)
    return out.getvalue()


def _finalize(wb: openpyxl.Workbook) -> bytes:
    """Pin the doc timestamps + the ZIP entry times -> deterministic `.xlsx` bytes."""
    wb.properties.created = _FIXED
    wb.properties.modified = _FIXED
    buf = io.BytesIO()
    wb.save(buf)
    return _normalize_zip(buf.getvalue())


# --------------------------------------------------------------------------- #
# Each builder returns the exact committed bytes for one fixture. The docstring
# pins the EXPECTED adapter routing (claims vs unextracted) — the golden test
# in test_excel_golden.py asserts these against the LIVE adapter.
# --------------------------------------------------------------------------- #


def _formula_with_cache_bytes() -> bytes:
    """A formula cell that DOES carry a cached value -> the adapter emits the cache as a claim.

    Layout: ``Calc!A1=2`` (literal), ``Calc!A2='=A1*10'`` (formula).
    Routing: A1 -> claim "2"; A2 -> claim "20" (the injected cache, value->string of int 20).
             #claims=2, #unextracted=0.

    Technique: openpyxl writes the formula but NO cache (`<f>A1*10</f><v />`). We rewrite the
    worksheet XML, replacing the empty `<v />` with `<v>20</v>`, so the DATA view (`data_only=True`)
    reads 20. The zip is rebuilt copying every entry in its original order (deterministic).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = 2
    ws["A2"] = "=A1*10"
    raw = _finalize(wb)

    sheet_xml = "xl/worksheets/sheet1.xml"
    zin = zipfile.ZipFile(io.BytesIO(raw))
    xml = zin.read(sheet_xml).decode("utf-8")
    # openpyxl serializes the uncached formula cell as `<f>A1*10</f><v />`; inject the cache.
    patched = xml.replace("<f>A1*10</f><v />", "<f>A1*10</f><v>20</v>")
    assert patched != xml, "formula-cache injection anchor not found (openpyxl XML shape changed)"

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():  # original entry order -> deterministic
            data = zin.read(item.filename)
            if item.filename == sheet_xml:
                data = patched.encode("utf-8")
            zout.writestr(item, data)
    return out.getvalue()


def _formula_no_cache_bytes() -> bytes:
    """The FAITHFULNESS CRUX: a formula cell with NO cache (the natural openpyxl output).

    Layout: ``Calc!A1=2`` (literal), ``Calc!A2='=A1*10'`` (formula, no cache).
    Routing: A1 -> claim "2"; A2 -> unextracted (formula-no-cache) — NEVER a "0"/"" claim.
             #claims=1, #unextracted=1.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws["A1"] = 2
    ws["A2"] = "=A1*10"  # openpyxl writes no cache -> data_only=True reads None -> unextracted
    return _finalize(wb)


def _merged_cells_bytes() -> bytes:
    """A merged range whose anchor value is emitted ONCE (the covered cells are blank, not drops).

    Layout: merge ``A1:B2`` with anchor ``A1='merged header'``; ``A3='below'``.
    Routing: A1 -> claim "merged header" (once); the 3 covered cells are None -> skipped (NOT
             drops); A3 -> claim "below". #claims=2, #unextracted=0.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws["A1"] = "merged header"
    ws.merge_cells("A1:B2")
    ws["A3"] = "below"
    return _finalize(wb)


def _multi_sheet_bytes() -> bytes:
    """Two sheets with values on each, to pin workbook-order, row-major traversal.

    Layout: sheet "First" (A1="first one", A2="first two"); sheet "Second" (A1="second one").
    Routing: 3 claims in workbook order ["first one", "first two", "second one"].
             #claims=3, #unextracted=0.
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1["A1"] = "first one"
    ws1["A2"] = "first two"
    ws2 = wb.create_sheet("Second")
    ws2["A1"] = "second one"
    return _finalize(wb)


def _mixed_types_bytes() -> bytes:
    """One cell per value->string rule: text, int, float, bool, date, datetime.

    Layout: A1=text, A2=int 42, A3=float 3.14, A4=bool True, A5=date, A6=datetime.
    Routing (value->string): "hello text", "42", "3.14", "TRUE", "2026-06-17T00:00:00"
            (openpyxl returns a date cell as a datetime), "2026-06-17T09:30:00".
            #claims=6, #unextracted=0.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws["A1"] = "hello text"
    ws["A2"] = 42
    ws["A3"] = 3.14
    ws["A4"] = True
    ws["A5"] = date(2026, 6, 17)
    ws["A6"] = datetime(2026, 6, 17, 9, 30, 0)
    return _finalize(wb)


def _empty_cells_bytes() -> bytes:
    """Sparse cells with genuine blanks between them (blank != formula-no-cache).

    Layout: A1="top"; A2 blank; A3 blank; A4="bottom" (B-column untouched).
    Routing: blanks are skipped and are NOT drops. #claims=2 ["top","bottom"], #unextracted=0.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sparse"
    ws["A1"] = "top"
    ws["A4"] = "bottom"
    return _finalize(wb)


def _error_cell_bytes() -> bytes:
    """A cell whose data_type is 'e' (an Excel error) -> unextracted, never a claim.

    Layout: A1="real value" (literal), A2=#DIV/0! (an error cell, data_type 'e').
    Routing: A1 -> claim "real value"; A2 -> unextracted (error cell). #claims=1, #unextracted=1.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Err"
    ws["A1"] = "real value"
    ws["A2"] = "#DIV/0!"  # openpyxl recognizes the error literal -> data_type 'e'
    return _finalize(wb)


def _chart_or_image_bytes() -> bytes:
    """A sheet carrying a CHART -> the silent-loss disclosure (chart content is NEVER extracted).

    Layout: a tiny data table (A1="cat",A2="x",B1="val",B2=5) + a BarChart anchored at D2.
    Routing: the 4 data cells -> 4 claims ["cat","x","val","5"]; the chart -> 1 unextracted
             disclosure (chart content out of scope). #claims=4, #unextracted=1.

    (Pillow is not a declared dependency, so a real image cannot be embedded. A chart exercises
    the identical `ws._charts` disclosure path the adapter walks for charts AND images.)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Charted"
    ws["A1"] = "cat"
    ws["A2"] = "x"
    ws["B1"] = "val"
    ws["B2"] = 5
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=2), titles_from_data=True)
    ws.add_chart(chart, "D2")
    return _finalize(wb)


# fixture filename -> the bytes-builder for it. Keys define the committed corpus.
FIXTURES: dict[str, "callable[[], bytes]"] = {  # type: ignore[type-arg]
    "formula_with_cache.xlsx": _formula_with_cache_bytes,
    "formula_no_cache.xlsx": _formula_no_cache_bytes,
    "merged_cells.xlsx": _merged_cells_bytes,
    "multi_sheet.xlsx": _multi_sheet_bytes,
    "mixed_types.xlsx": _mixed_types_bytes,
    "empty_cells.xlsx": _empty_cells_bytes,
    "error_cell.xlsx": _error_cell_bytes,
    "chart_or_image.xlsx": _chart_or_image_bytes,
}


def main() -> None:
    for name, builder in FIXTURES.items():
        data = builder()
        (HERE / name).write_bytes(data)
        print(f"wrote {name} ({len(data)} bytes)")


if __name__ == "__main__":
    main()
