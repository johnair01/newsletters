"""R4 Wave-0 probe — pin the openpyxl chart/image DETECTION attribute names (assumption A1).

CONTEXT R4 / RESEARCH assumption A1: the Excel adapter (Plan 05-03) must DISCLOSE charts and
images it cannot extract (their CONTENT is a deferred idea — out of scope) by routing their mere
presence to ``unextracted[]``. To do that it needs the attribute openpyxl uses to surface parsed
drawings on a worksheet. Research ASSUMES ``ws._charts`` / ``ws._images`` (a LOW-confidence
tertiary source). This probe CONFIRMS those names against the INSTALLED openpyxl so Plan 05-03
relies on a verified attribute, not an assumption — and FAILS LOUDLY (surfacing the real names)
if a future openpyxl version renames them.

The probe authors a tiny in-memory workbook (a worksheet with a BarChart), saves to a BytesIO,
reloads in standard mode, and asserts the worksheet exposes the chart via ``ws._charts`` and that
``ws._images`` exists as the image-detection attribute. It is skipped cleanly when the ``[excel]``
extra is not installed, so the bare-install gate is unaffected; it MUST run and pass in the dev
.venv where openpyxl is present.

CONFIRMED against openpyxl 3.1.5 (run 2026-06-17):
- ``ws._charts``  — list of chart objects; a re-loaded workbook KEEPS charts (BarChart survives).
- ``ws._images``  — list of image objects; PRESENT (attribute exists) but openpyxl LOSES images
  from existing files on reload, so a reloaded ws._images is typically empty even if the source
  had images. Detection of images is therefore best done on the freshly-built/parsed object;
  the attribute itself is the correct detection surface.
- Both attributes exist ONLY in standard mode (``read_only=False``). In ``read_only=True`` the
  worksheet has NEITHER ``_charts`` nor ``_images`` — a second reason Plan 05-03 must load with
  ``read_only=False`` (already mandated by R3 for merged-cell faithfulness).
"""

from __future__ import annotations

import io

import pytest

# Skip the whole module cleanly on a bare install (no [excel]) so the bare-install gate is
# unaffected. In the dev/CI [excel] env this import succeeds and the probe runs.
openpyxl = pytest.importorskip("openpyxl", reason="[excel] extra not installed; probe runs in dev/CI")


def _build_workbook_with_chart() -> io.BytesIO:
    """Author a tiny in-memory .xlsx whose first sheet carries a BarChart, return its bytes."""
    from openpyxl.chart import BarChart, Reference

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Probe"
    for i, value in enumerate([1, 2, 3, 4], start=1):
        ws.cell(row=i, column=1, value=value)
    data = Reference(ws, min_col=1, min_row=1, max_row=4)
    chart = BarChart()
    chart.add_data(data)
    ws.add_chart(chart, "C1")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_worksheet_exposes_charts_attribute() -> None:
    """``ws._charts`` is the chart-detection attribute, and a re-loaded chart is discoverable.

    Pins assumption A1 for charts: if openpyxl ever renames ``_charts`` this fails LOUDLY,
    surfacing the worksheet's actual drawing-related attributes for Plan 05-03 to adopt.
    """
    buf = _build_workbook_with_chart()
    wb = openpyxl.load_workbook(buf, data_only=False, read_only=False)
    ws = wb["Probe"]
    try:
        assert hasattr(ws, "_charts"), (
            "openpyxl worksheet has NO '_charts' attribute — assumption A1 is WRONG for this "
            "version. Drawing-related attrs present: "
            f"{sorted(a for a in dir(ws) if 'chart' in a.lower() or 'draw' in a.lower())}"
        )
        # A standard-mode reload KEEPS charts (verified: BarChart survives round-trip).
        assert len(ws._charts) == 1, (
            f"expected exactly 1 chart after reload, found {len(ws._charts)}: "
            f"{[type(c).__name__ for c in ws._charts]}"
        )
        assert type(ws._charts[0]).__name__ == "BarChart", type(ws._charts[0]).__name__
    finally:
        wb.close()


def test_worksheet_exposes_images_attribute() -> None:
    """``ws._images`` is the image-detection attribute (present as a list in standard mode).

    Pins assumption A1 for images. NOTE: openpyxl loses images from EXISTING files on reload, so
    a reloaded ``_images`` is typically empty; the attribute's PRESENCE (the detection surface) is
    what Plan 05-03 relies on, applied to the freshly-parsed worksheet object.
    """
    buf = _build_workbook_with_chart()  # no image, but the attribute must still exist
    wb = openpyxl.load_workbook(buf, data_only=False, read_only=False)
    ws = wb["Probe"]
    try:
        assert hasattr(ws, "_images"), (
            "openpyxl worksheet has NO '_images' attribute — assumption A1 is WRONG for this "
            "version. Image-related attrs present: "
            f"{sorted(a for a in dir(ws) if 'image' in a.lower() or 'draw' in a.lower())}"
        )
        # It is a list-like collection (empty here — no image authored / reload drops images).
        assert list(ws._images) == [], f"unexpected images on a chart-only sheet: {ws._images}"
    finally:
        wb.close()


def test_detection_attributes_absent_in_read_only_mode() -> None:
    """Document WHY Plan 05-03 must load with ``read_only=False``: read_only LOSES the attrs.

    In ``read_only=True`` the worksheet object exposes NEITHER ``_charts`` nor ``_images`` — so
    chart/image disclosure is impossible in read_only mode. This corroborates R3's standard-mode
    mandate (already required for faithful merged-cell accounting) with a second, independent
    reason and pins it as a committed expectation.
    """
    buf = _build_workbook_with_chart()
    wb = openpyxl.load_workbook(buf, read_only=True)
    try:
        ws = wb["Probe"]
        assert not hasattr(ws, "_charts"), (
            "read_only worksheet UNEXPECTEDLY exposes '_charts' — re-evaluate the load-mode "
            "rationale for Plan 05-03."
        )
        assert not hasattr(ws, "_images"), (
            "read_only worksheet UNEXPECTEDLY exposes '_images' — re-evaluate the load-mode "
            "rationale for Plan 05-03."
        )
    finally:
        wb.close()
