"""Cross-adapter determinism proof for the shared deterministic-timestamp helper (L1).

The point (CONTEXT decision 0 / research-locked L1): before Phase-6's PptxAdapter copies the
adapter timestamp pattern, lock that pattern to a DETERMINISTIC sentinel. Today email/excel source
``Source.timestamp`` from the document's intrinsic date (email ``Date`` header, xlsx docProps
``created``) but fall back to ``Source``'s ``_utcnow()`` default-factory wall-clock when absent — so
two parses of the SAME bytes (a real document with no intrinsic date) produce NON-equal Sources,
breaking the determinism / round-trip-parity property the whole adapter contract depends on.

Two layers:

1. **Helper-level tests** (Task 1, RED→GREEN) — ``deterministic_timestamp`` returns the intrinsic
   value (coercing tz-naive to UTC) or ``EPOCH_ZERO`` (never now()); ``EPOCH_ZERO`` is tz-aware UTC.

2. **Cross-adapter determinism proof** (Task 2) — an ``.eml`` with NO Date header and an ``.xlsx``
   with NO docProps ``created`` each parse to BYTE-IDENTICAL Sources twice (``model_dump_json``
   equal) with ``timestamp == EPOCH_ZERO``; an intrinsic-timestamp doc still uses its own date.
   Parametrized across email + excel; the PptxAdapter param joins this exact matrix in Plan 06-04
   (the SEAM is marked below — do NOT import python-pptx here).
"""

from __future__ import annotations

import importlib.util
import io
from datetime import datetime, timezone
from email.message import EmailMessage

import pytest

from newsletters.adapters._timestamps import EPOCH_ZERO, deterministic_timestamp
from newsletters.adapters.email_adapter import EmailAdapter

# ============================================================================ #
# Task 1 — helper-level RED tests (the sentinel + the coercion rules)
# ============================================================================ #


def test_epoch_zero_is_tz_aware_utc_1970() -> None:
    """EPOCH_ZERO is the obvious-to-a-reviewer 1970-01-01, tz-aware UTC (importable for asserts)."""
    assert EPOCH_ZERO == datetime(1970, 1, 1, tzinfo=timezone.utc)
    assert EPOCH_ZERO.tzinfo is not None
    assert EPOCH_ZERO.utcoffset() == timezone.utc.utcoffset(None)


def test_deterministic_timestamp_none_returns_epoch_zero() -> None:
    """No intrinsic timestamp -> EPOCH_ZERO, NEVER now()/wall-clock (the determinism fix)."""
    assert deterministic_timestamp(None) == EPOCH_ZERO


def test_deterministic_timestamp_aware_returns_unchanged() -> None:
    """An already-tz-aware intrinsic timestamp is preserved verbatim (no regression)."""
    aware = datetime(2026, 6, 17, 9, 0, 0, tzinfo=timezone.utc)
    assert deterministic_timestamp(aware) == aware
    assert deterministic_timestamp(aware).tzinfo is not None


def test_deterministic_timestamp_naive_coerced_to_utc() -> None:
    """A tz-naive intrinsic timestamp is coerced to UTC via replace(tzinfo=utc) (deterministic)."""
    naive = datetime(2026, 6, 17, 9, 0, 0)
    coerced = deterministic_timestamp(naive)
    assert coerced == datetime(2026, 6, 17, 9, 0, 0, tzinfo=timezone.utc)
    assert coerced.tzinfo is not None


def test_deterministic_timestamp_is_pure_and_repeatable() -> None:
    """Same input -> same output across repeated calls (no hidden now()/state)."""
    assert deterministic_timestamp(None) == deterministic_timestamp(None)
    aware = datetime(2020, 1, 2, 3, 4, 5, tzinfo=timezone.utc)
    assert deterministic_timestamp(aware) == deterministic_timestamp(aware)


# ============================================================================ #
# Task 2 — cross-adapter determinism proof (parametrized across email + excel)
# ============================================================================ #
#
# Each case is a self-contained recipe so the PptxAdapter joins by appending ONE entry in Plan
# 06-04 (the seam): factory() -> a fresh adapter; no_intrinsic_bytes() -> bytes of a document with
# NO intrinsic timestamp; with_intrinsic() -> (bytes, expected_datetime) for a document that HAS an
# intrinsic timestamp. The PPTX entry will use core_properties.created (or None) — do NOT import
# python-pptx in this file; that dependency is gated behind the [pptx] extra in Plan 06-02.
# --------------------------------------------------------------------------- #

_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None


def _eml_no_date() -> bytes:
    """An .eml with a body but NO Date header -> the adapter has no intrinsic timestamp."""
    msg = EmailMessage()
    msg["From"] = "a@example.com"
    msg["To"] = "b@example.com"
    msg["Subject"] = "no date here"
    msg.set_content("A body paragraph.\n\nAnother paragraph.")
    return msg.as_bytes()


def _eml_with_date() -> tuple[bytes, datetime]:
    """An .eml WITH a Date header -> the adapter must keep that intrinsic timestamp (no regression)."""
    msg = EmailMessage()
    msg["From"] = "a@example.com"
    msg["To"] = "b@example.com"
    msg["Subject"] = "dated"
    msg["Date"] = "Wed, 17 Jun 2026 09:00:00 +0000"
    msg.set_content("Body.")
    return msg.as_bytes(), datetime(2026, 6, 17, 9, 0, 0, tzinfo=timezone.utc)


def _email_factory():
    return EmailAdapter()


def _strip_created_from_xlsx(raw: bytes) -> bytes:
    """Rewrite an .xlsx zip with the docProps `created`/`modified` elements removed.

    openpyxl CANNOT save a workbook with `properties.created = None` (it calls `.set()` on it during
    serialization) AND it fabricates a wall-clock `created` (`created or now()`) on save whenever
    absent — so there is no high-level way to author a deterministic "no intrinsic timestamp" .xlsx.
    We therefore strip the elements from the saved zip directly (stdlib only). The adapter's
    `intrinsic_created` reads the raw XML, so it sees a genuinely-absent `created` -> None -> EPOCH_ZERO.
    """
    import re
    import zipfile

    src = zipfile.ZipFile(io.BytesIO(raw))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "docProps/core.xml":
                data = re.sub(rb"<dcterms:created[^>]*>[^<]*</dcterms:created>", b"", data)
                data = re.sub(rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>", b"", data)
            zf.writestr(item, data)
    return out.getvalue()


def _xlsx_no_created() -> bytes:
    """An .xlsx with NO docProps `created` element -> no intrinsic timestamp (deterministic)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "text"
    ws["A2"] = 42
    buf = io.BytesIO()
    wb.save(buf)
    return _strip_created_from_xlsx(buf.getvalue())


def _xlsx_with_created() -> tuple[bytes, datetime]:
    """An .xlsx WITH docProps `created` -> the adapter must keep that intrinsic timestamp."""
    import openpyxl

    created = datetime(2026, 6, 17, 9, 0, 0)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "text"
    wb.properties.created = created
    buf = io.BytesIO()
    wb.save(buf)
    # openpyxl serializes `created` as a tz-naive UTC W3CDTF instant; intrinsic_created reads it from
    # the raw XML and coerces it to UTC, so the expected Source.timestamp is the UTC-aware value.
    return buf.getvalue(), created.replace(tzinfo=timezone.utc)


def _excel_factory():
    from newsletters.adapters.excel_adapter import ExcelAdapter

    return ExcelAdapter()


# (id, adapter_factory, no_intrinsic_bytes_loader, with_intrinsic_loader)
_DETERMINISM_CASES = [
    pytest.param(_email_factory, _eml_no_date, _eml_with_date, id="email"),
    pytest.param(
        _excel_factory,
        _xlsx_no_created,
        _xlsx_with_created,
        id="excel",
        marks=pytest.mark.skipif(
            not _HAS_OPENPYXL, reason="optional [excel] extra (openpyxl) not installed"
        ),
    ),
    # SEAM: the PptxAdapter param joins HERE in Plan 06-04 — a (_pptx_factory,
    # _pptx_no_created, _pptx_with_created) entry using core_properties.created. Do NOT import
    # python-pptx in this file; it lives behind the [pptx] extra (Plan 06-02).
]


@pytest.mark.parametrize("factory,no_intrinsic_bytes,_with_intrinsic", _DETERMINISM_CASES)
def test_no_intrinsic_timestamp_parses_byte_identical_twice(
    factory, no_intrinsic_bytes, _with_intrinsic
) -> None:
    """A doc with NO intrinsic timestamp parses to byte-identical Sources twice (timestamp==EPOCH_ZERO).

    This is the property the front-fix exists for: with the old now() fallback, two parses of the
    SAME bytes produced different timestamps and so non-equal Sources.
    """
    raw = no_intrinsic_bytes()
    path = "no-intrinsic.doc"
    s1, _u1, _x1 = factory().parse(raw, path)
    s2, _u2, _x2 = factory().parse(raw, path)

    assert s1.timestamp == EPOCH_ZERO, (
        "a document with no intrinsic timestamp must use the deterministic EPOCH_ZERO sentinel"
    )
    assert s1.model_dump_json() == s2.model_dump_json(), (
        "two parses of identical bytes must yield byte-identical Sources (determinism)"
    )


@pytest.mark.parametrize("factory,_no_intrinsic_bytes,with_intrinsic", _DETERMINISM_CASES)
def test_intrinsic_timestamp_is_preserved(
    factory, _no_intrinsic_bytes, with_intrinsic
) -> None:
    """A doc WITH an intrinsic timestamp still uses that date (no regression from the front-fix)."""
    raw, expected = with_intrinsic()
    source, _u, _x = factory().parse(raw, "with-intrinsic.doc")
    assert source.timestamp == expected, (
        "an intrinsic document date must be preserved, never replaced by EPOCH_ZERO"
    )
    assert source.timestamp != EPOCH_ZERO
